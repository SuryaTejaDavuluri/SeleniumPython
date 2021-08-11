import pytest
import openpyxl
from Pages.RadioButtonsDemoPage import RadioButtonPage
from Tests.test_base import BaseTest


class TestRadioButtons(BaseTest):

    def test_close_popup(self):
        self.radiobuttons = RadioButtonPage(self.driver)
        self.radiobuttons.close_pop_up()

    def get_data(sheetname):
        workbook = openpyxl.load_workbook('C:/Users/RC08508/PycharmProjects/SeleniumPython/Code/TestData/testdata.xlsx')
        sheet = workbook[sheetname]
        row_count = sheet.max_row
        col_count = sheet.max_column
        mainList = []
        for i in range(2, row_count + 1):
            dataList = []
            for j in range(1, col_count + 1):
                data = sheet.cell(row=i, column=j).value
                dataList.insert(j, data)
            mainList.insert(i, dataList)
        return mainList

    @pytest.mark.parametrize('value', get_data("RadioButton1"))
    def test_single_radio_button(self, value):
        self.radiobuttons = RadioButtonPage(self.driver)
        self.radiobuttons.get_page_title()
        self.radiobuttons.get_menu()
        self.radiobuttons.get_page()
        print(value)
        # self.radiobuttons.chosen_gender(value)

    # @pytest.mark.parametrize('value', ["Male"])
    # def test_single_radio_button(self, value):
    #     self.radiobuttons = RadioButtonPage(self.driver)
    #     self.radiobuttons.get_page_title()
    #     self.radiobuttons.get_menu()
    #     self.radiobuttons.get_page()
    #     self.radiobuttons.chosen_gender(value)
