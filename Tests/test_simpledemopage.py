import pytest
import openpyxl
from Pages.SimpleDemoPage import SimpleDemoPage
from Tests.test_base import BaseTest
import time


class TestSimpleForm(BaseTest):

    def test_close_popup(self):
        self.simpleDemoPage = SimpleDemoPage(self.driver)
        self.simpleDemoPage.close_pop_up()

    def get_data(sheetname):
        workbook = openpyxl.load_workbook('C:/Users/RC08508/PycharmProjects/SeleniumPython/Code/TestData/testdata.xlsx')
        sheet = workbook[sheetname]
        row_count = sheet.max_row
        col_count = sheet.max_column
        mainList = []
        for i in range(2, row_count + 1):
            dataList = []
            for j in range(1, col_count + 1):
                data = sheet.cell(row=i, column=j).value
                dataList.insert(j, data)
            mainList.insert(i, dataList)
        return mainList

    @pytest.mark.parametrize('value', get_data("SimpleDemo1"))
    def test_single_input_field(self, value):
        self.simpleDemoPage = SimpleDemoPage(self.driver)
        self.simpleDemoPage.get_page_title()
        self.simpleDemoPage.get_menu()
        self.simpleDemoPage.get_page()
        self.simpleDemoPage.enter_message(value)
        # time.sleep(5)
        self.simpleDemoPage.get_buttons()
        x = self.simpleDemoPage.show_message()
        assert [x] == value

    @pytest.mark.parametrize("value1, value2", get_data("SimpleDemo2"))
    def test_multi_input_field(self, value1, value2):
        self.simpleDemoPage = SimpleDemoPage(self.driver)
        self.simpleDemoPage.get_page_title()
        self.simpleDemoPage.get_menu()
        self.simpleDemoPage.get_page()
        self.simpleDemoPage.enter_sum1(value1)
        self.simpleDemoPage.enter_sum2(value2)
        self.simpleDemoPage.calc_sum()
        x = self.simpleDemoPage.show_sum_value()
        assert int(x) == int(value1) + int(value2)


