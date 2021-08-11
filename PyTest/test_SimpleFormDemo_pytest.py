from selenium.webdriver.common.by import By
# from selenium.webdriver.support.wait import WebDriverWait
# from selenium.webdriver.support import expected_conditions as ec
import pytest
import xlrd
import time
import openpyxl

def get_data(sheet):
    workbook = openpyxl.load_workbook('C:/Users/RC08508/PycharmProjects/SeleniumPython/Code/TestData/testdata.xlsx')
    sheet = workbook[sheet]
    row_count = sheet.max_row
    col_count = sheet.max_column
    mainList = []
    for i in range(2, row_count+1):
        dataList = []
        for j in range(1, col_count+1):
            data = sheet.cell(row=i, column=j).value
            dataList.insert(j, data)
        mainList.insert(i, dataList)
    return mainList


@pytest.mark.usefixtures('init_driver')
class BaseTest:
    pass


class TestSimpleForm(BaseTest):
    @pytest.mark.parametrize('value', get_data("SimpleDemo1"))
    def test_single_input_field(self, value):
        self.driver.find_element(By.LINK_TEXT, 'Input Forms').click()
        self.driver.find_element(By.LINK_TEXT, 'Simple Form Demo').click()
        buttons = self.driver.find_elements(By.XPATH, '//button[@type="button" and @class="btn btn-default"]')
        print(buttons)
        self.driver.find_element(By.ID, 'user-message').clear()
        self.driver.find_element(By.ID, 'user-message').send_keys(value)
        for i in buttons:
            # print(i.text)
            if i.text == 'Show Message':
                i.click()
        x = self.driver.find_element(By.ID, 'display').text
        print(x)
        print(value)
        if [x] == value:
            print('Text Matched')
        else:
            print("Text MisMatched")

    @pytest.mark.parametrize("value1, value2", get_data("SimpleDemo2"))
    def test_multi_input_field(self, value1, value2):
        buttons = self.driver.find_elements(By.XPATH, '//button[@type="button" and @class="btn btn-default"]')
        self.driver.find_element(By.ID, 'sum1').clear()
        self.driver.find_element(By.ID, 'sum1').send_keys(value1)
        self.driver.find_element(By.ID, 'sum2').clear()
        self.driver.find_element(By.ID, 'sum2').send_keys(value2)
        for j in buttons:
            if j.text == 'Get Total':
                j.click()
        y = int(self.driver.find_element(By.ID, 'displayvalue').text)
        print(y)
        print(value1)
        print(value2)
        if y == int(value1) + int(value2):
            print('Value Matched')
        else:
            print("Value MisMatched")


